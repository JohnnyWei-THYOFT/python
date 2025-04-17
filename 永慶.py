from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException
import time
import openpyxl


#創建xlsx
wb = openpyxl.Workbook()
ws = wb.active

ws['A1']="地段" 
ws['B1']="價格" 
ws['C1']="樓層" 
ws['D1']="坪數" 
ws['E1']="格局"


options = webdriver.ChromeOptions()
options.add_argument("--disable-blink-features=AutomationControlled") 
prefs = {"profile.default_content_setting_values": {"notifications": 2}}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(service=Service(excutable_path=ChromeDriverManager().install(),options=options))
driver.get("https://rent.yungching.com.tw/")

def rent(x,y):
    driver.find_element(By.CSS_SELECTOR,'span[title='+y+']').click()
    time.sleep(2)
    driver.find_element(By.ID,x).click()
    time.sleep(2)
    while True:
        try:
            last_height = driver.execute_script("return document.body.scrollHeight")
            time.sleep(2)  # 等待一段時間讓頁面加載新的內容
            new_height = driver.execute_script("return document.body.scrollHeight")
       
            if new_height == last_height:
                last_height = new_height
        
            houseboxes = driver.find_elements(By.CLASS_NAME, 'housebox')
            for i in houseboxes:
                a=i.find_elements(By.CSS_SELECTOR,'ul.houseul01 li')[0].text
                b=i.find_element(By.CLASS_NAME,'price').text
                c=i.find_elements(By.CSS_SELECTOR,'ul.houseul03 li')[0].text
                d=i.find_elements(By.CSS_SELECTOR,'ul.houseul01 li')[1].text
                e=i.find_elements(By.CSS_SELECTOR,'ul.houseul03 li')[1].text
                print(a)
                print(b)
                print(c)
                print(d)
                print(e)
                ws.append([a,b,c,d,e])	
            print("---------")
    
            driver.find_element(By.CSS_SELECTOR,'a[class="next"]').click()  
            time.sleep(1)
        
        except ElementNotInteractableException:
            # 如果遇到不可交互的元素，跳出迴圈
            break
rent_type = ["Purpose01","Purpose02","Purpose03","Purpose04"]
city = ["台北市","新北市","桃園市"]
for j in city:
    for i in rent_type:
        rent(i,j)
        
wb.save("北北桃租屋.xlsx")



