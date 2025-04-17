import requests
from bs4 import BeautifulSoup
import openpyxl

#創建xlsx
wb = openpyxl.Workbook()
ws = wb.active

ws['A1']="職缺名稱" 
ws['B1']="職缺連結" 
ws['C1']="公司名稱" 
ws['D1']="工作地區" 
ws['E1']="薪資待遇"


res=requests.get("https://www.104.com.tw/jobs/search/?jobsource=m_joblist_search&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&order=15&area=&jobcat=&page=1")
soup=BeautifulSoup(res.text)
page=1
while soup.find_all("div",class_="info-container") != []:
    print("===============================")
    print("正在抓取",page,"頁")
    print("===============================")
    for job in soup.find_all("div",class_="info-container"):
        print(job.a.text)
        print(job.a["href"])
        print(job.find("div",class_="info-company mb-1").a.text)
        print(job.find("span",class_="info-tags__text font-weight-bold").text)
        print(job.find_all("span",class_="info-tags__text font-weight-bold")[3].text)
        print("----------------")
        ws.append([job.a.text,job.a["href"],job.find("div",class_="info-company mb-1").a.text,job.find("span",class_="info-tags__text font-weight-bold").text,job.find_all("span",class_="info-tags__text font-weight-bold")[3].text])
    page+=1    
    res=requests.get("https://www.104.com.tw/jobs/search/?jobsource=m_joblist_search&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&order=15&area=&jobcat=&page="+str(page))
    soup=BeautifulSoup(res.text)
    wb.save("104職缺.xlsx")
