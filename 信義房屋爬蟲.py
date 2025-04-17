from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import openpyxl

#創建xlsx
wb = openpyxl.Workbook()
ws = wb.active

ws['A1']="地段" 
ws['B1']="價格" 
ws['C1']="樓層" 
ws['D1']="坪數" 
ws['E1']="格局"

city = ["Taipei-city","NewTaipei-city","Taoyuan-city","Keelung-city"]

options = webdriver.ChromeOptions()
options.add_argument("--disable-blink-features=AutomationControlled")  # 隱藏自動化特徵
prefs = {"profile.default_content_setting_values": {"notifications": 2}}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(service=Service(excutable_path=ChromeDriverManager().install(),options=options))

for j in city:
    page=1
    driver.get("https://www.sinyi.com.tw/rent/list/"+j+"/house-rent1-rent2-rent3-use/"+str(page)+".html")
    while driver.find_elements(By.CLASS_NAME,"item_detailbox") != []:
        
        x = driver.find_elements(By.CLASS_NAME,"item_detailbox")
        print("第"+str(page)+"頁")
        for i in x:
            z = []
            a = (i.find_element(By.CSS_SELECTOR,'span[class="num num-text"]').text)
            a = a.split(" / ")
            if len(a) > 2:
            # 如果有兩個或以上的 "/"
                z=a[2]  # 只取第二個 "/" 後的部分
            elif len(a) > 1:
            # 如果只有一個 "/"
                z=a[1]  # 只取 "/" 後的部分
            else:
            # 如果沒有 "/"
                z=a[0]  # 保留原始資料
            b = (i.find_elements(By.CSS_SELECTOR,'span[class="num"]')[3].text)
            c = (i.find_elements(By.CSS_SELECTOR,'span[class="num"]')[1].text)
            d = (i.find_elements(By.CSS_SELECTOR,'span[class="num"]')[0].text)
            e = (i.find_elements(By.CSS_SELECTOR,'span[class="num"]')[2].text)
            ws.append([z,b,c,d,e])
            
        page+=1
        driver.get("https://www.sinyi.com.tw/rent/list/"+j+"/house-rent1-rent2-rent3-use/"+str(page)+".html")
wb.save("北北基桃租屋.xlsx")
print(j)

